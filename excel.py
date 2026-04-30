"""Excel 填写：根据日志数据生成事件上报 Excel"""

import os
from datetime import datetime
from copy import copy

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    TEMPLATE_XLSX, ORG_NAME,
    ATTACK_CLASS_MAP, DEFAULT_CLASS,
    translate_attack_type, translate_risk_level,
)

HEADERS_XLSX = [
    "序号", "事件标题", "事件描述", "风险级别",
    "攻击IP", "攻击端口", "攻击IP所属国家", "攻击IP所属省市",
    "目地IP", "目地端口", "目的IP所属国家", "目的IP所属省市",
    "攻击类型", "攻击协议", "告警时间", "重保工单(是/否)",
    "防护拦截设备名称", "是否设备自动拦截(是/否)", "被攻击系统名称",
    "事件分类", "分类子类", "payload", "事件分类补充", "分类子类补充"
]

COL_WIDTHS = [6, 30, 50, 8, 18, 10, 16, 12, 18, 10, 16, 12,
              14, 10, 22, 14, 20, 22, 28, 22, 22, 35, 16, 16]


def get_event_class(attack_type):
    for key, val in ATTACK_CLASS_MAP.items():
        if key in attack_type:
            return val
    return DEFAULT_CLASS


def get_protocol(log_item):
    scheme = (log_item.get("scheme") or "").upper()
    if scheme:
        return scheme
    return {443: "HTTPS", 80: "HTTP"}.get(log_item.get("dst_port"), "TCP")


def parse_timestamp(log_item):
    for key in ("human_timestamp", "human_req_start_time"):
        val = log_item.get(key)
        if val:
            try:
                return datetime.strptime(val[:19], "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                pass
    ts = log_item.get("timestamp") or log_item.get("req_start_time")
    if ts:
        try:
            return datetime.fromtimestamp(int(ts)).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            pass
    return ""


def build_description(log_item, attack_type, system_name):
    reason = log_item.get("req_reason") or log_item.get("reason") or ""
    src_ip = log_item.get("src_ip", "")
    attack_text = f"受到{attack_type}攻击" if attack_type != "未知" else "受到攻击"
    return (f"{ORG_NAME}{system_name}{attack_text}，{reason}，"
            f"此次攻击已被WAF自动拦截，未造成实际影响，建议封堵该IP（{src_ip}）。")


def build_row_data(log_item, seq):
    attack_type_raw = log_item.get("attack_type")
    if not attack_type_raw or attack_type_raw == "None":
        attack_type_raw = log_item.get("req_attack_type") or "未知"
    attack_type = translate_attack_type(attack_type_raw)
    src_province = log_item.get("province") or "未知"
    website = log_item.get("website_name") or log_item.get("host") or ""
    system_name = website.split("-")[0] if "-" in website else website
    event_class, sub_class = get_event_class(attack_type)

    risk_level = log_item.get("risk_level") or log_item.get("req_risk_level") or ""
    risk_level = translate_risk_level(risk_level)

    return {
        "序号": seq,
        "事件标题": f"{ORG_NAME}受到{attack_type}攻击" if attack_type != "未知" else f"{ORG_NAME}受到攻击",
        "事件描述": build_description(log_item, attack_type, system_name),
        "风险级别": risk_level,
        "攻击IP": log_item.get("src_ip", ""),
        "攻击端口": log_item.get("src_port", ""),
        "攻击IP所属国家": log_item.get("country", ""),
        "攻击IP所属省市": src_province,
        "目地IP": log_item.get("dst_ip", ""),
        "目地端口": log_item.get("dst_port", ""),
        "目的IP所属国家": "中国",
        "目的IP所属省市": "北京",
        "攻击类型": attack_type,
        "攻击协议": get_protocol(log_item),
        "告警时间": parse_timestamp(log_item),
        "重保工单(是/否)": "否",
        "防护拦截设备名称": "WAF",
        "是否设备自动拦截(是/否)": "是",
        "被攻击系统名称": system_name,
        "事件分类": event_class,
        "分类子类": sub_class,
        "payload": str(log_item.get("payload") or log_item.get("url_path") or ""),
        "事件分类补充": "",
        "分类子类补充": "",
    }


def make_header_style():
    thin = Side(style="thin", color="AAAAAA")
    return (
        Font(name="Arial", bold=True, color="FFFFFF", size=10),
        PatternFill("solid", start_color="2E4057"),
        Alignment(horizontal="center", vertical="center", wrap_text=True),
        Border(left=thin, right=thin, top=thin, bottom=thin),
    )


def make_data_style():
    thin = Side(style="thin", color="DDDDDD")
    return (
        Font(name="Arial", size=9),
        Alignment(horizontal="left", vertical="center", wrap_text=True),
        Border(left=thin, right=thin, top=thin, bottom=thin),
    )


def fill_excel(log_item, seq, xlsx_out):
    """根据模板填写 Excel"""
    row_data = build_row_data(log_item, seq)

    if os.path.exists(TEMPLATE_XLSX):
        wb = load_workbook(TEMPLATE_XLSX)
        ws = wb["事件工单模板"]
        use_template = True
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "事件工单模板"
        hfont, hfill, halign, hborder = make_header_style()
        for col_idx, (header, width) in enumerate(zip(HEADERS_XLSX, COL_WIDTHS), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = halign
            cell.border = hborder
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"
        use_template = False

    # 写入数据行
    target_row = 2
    dfont, dalign, dborder = make_data_style()

    for col_idx in range(1, 25):
        header_name = ws.cell(row=1, column=col_idx).value
        dst_cell = ws.cell(row=target_row, column=col_idx)

        if use_template:
            src_cell = ws.cell(row=2, column=col_idx)
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.border = copy(src_cell.border)
                dst_cell.alignment = copy(src_cell.alignment)
        else:
            dst_cell.font = dfont
            dst_cell.alignment = dalign
            dst_cell.border = dborder

        if header_name and header_name in row_data:
            dst_cell.value = row_data[header_name]

    ws.row_dimensions[target_row].height = 45
    wb.save(xlsx_out)
